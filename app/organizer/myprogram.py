import numpy as np
import mojimoji
import openpyxl as px
import pandas as pd
from openpyxl.styles import Font
from openpyxl.styles.colors import Color
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
from openpyxl.styles import colors

from django.db.models.aggregates import Count
from django.db.models import Max
from django.core.exceptions import ObjectDoesNotExist

# Models
from competitions.models import Comp, Event, EventStatus, GR as GRecord
from organizer.models import Entry
from organizer.templatetags.organizer_tags import format_mark
from organizer.templatetags.organizer_filters import zen_to_han, sex_to_ja


"""
Default Styling
"""
# Color
color_default = Color(theme=1, tint=0.0)
# Font
font_default = Font(name='ＭＳ ゴシック',charset=128,family=3.0,b=False,i=False,strike=None,outline=None,shadow=None,condense=None,color=color_default,size=8,)
# Alignment
al_default = Alignment(shrinkToFit=None, textRotation=0, vertical='center', horizontal='center', indent=0.0, justifyLastLine=None, relativeIndent=0.0, wrapText=None, readingOrder=0.0)
# Border
border_default = Border(
    left=Side(border_style=None, color='FF000000'),
    right=Side(border_style=None, color='FF000000'),
    top=Side(border_style=None, color='FF000000'),
    bottom=Side(border_style=None, color='FF000000'),
    diagonal=Side(border_style=None, color='FF000000'),
    outline=Side(border_style=None, color='FF000000'),
    vertical=Side(border_style=None, color='FF000000'),
    horizontal=Side(border_style=None, color='FF000000')
)


class ProgramMaker:
    """
    Init
    """
    def __init__(self, comp, *args, **kwargs):
        self.comp = comp
        # Font
        self.font_small = Font(name='ＭＳ ゴシック',charset=128,size=6,)
        self.font_red = Font(name='ＭＳ ゴシック',charset=128,size=9,color=colors.RED)
        # Alignment
        self.al_wrap = Alignment(vertical='center', horizontal='center', wrapText=True)
        self.al_left = Alignment(vertical='center', horizontal='left')
        self.al_right = Alignment(vertical='center', horizontal='right')
        self.al_bottom = Alignment(vertical='bottom', horizontal='center')
        # Borde
        thin = Side(border_style="thin", color="000000")
        thick = Side(border_style="thick", color="000000")
        self.border_bottom = Border(top=None, left=None, right=None, bottom=thin)
        self.border_all = Border(top=thin, left=thin, right=thin, bottom=thin)
        self.border_all_thick = Border(top=thick, left=thick, right=thick, bottom=thick)
        event_status = EventStatus.objects.all()[0]
        self.entry_Null = Entry(event_status=event_status, bib="", name_family="", name_first="", kana_family="", kana_first="", grade="", club="", jaaf_branch="", personal_best="", sex="M", entry_status="Entry")

        # １ページに入る行数
        self.row_page_max = 56       # 高さ13では56行で1ページ
        # 前の組を書き込んだ行
        self.row_prev_group_top = 1
        self.row_prev_group_last = 1

        
    """
    Utils
    """
    def cell_posi(self, row, col):
        return col+str(row)


    def write_cell(self, cell, value, font=font_default, al=al_default, border=border_default):
        cell.value = value
        cell.font = font
        cell.alignment = al
        cell.border = border

    def format_name(self, entry):
        # 名前
        # できるだけ5文字に近くして戻す
        name_len  = len(entry.name_family)+len(entry.name_first)
        if name_len == 4:
            # 4文字
            return entry.name_family + u"\u3000" + entry.name_first
        elif  name_len < 4:
            # 4文字未満
            name = entry.name_family
            for i in range(0, 4 - name_len):
                name += u"\u3000"
            return (name + u"\u3000" + entry.name_first)
        else:
            # 5文字以上
            return entry.name_family + entry.name_first

    def format_kana(self, entry):
        # フリガナ
        kana = entry.kana_family + " " + entry.kana_first
        return mojimoji.zen_to_han(kana)


    def get_current_page(self, row):
        """
        現在書き込みが完了している行(row)が何ページ目の何行目かを返す.
        """
        # Param
        # - row: 書き始めの行番号
        
        #現在の場所を計算
        page_current = int((row-1)/self.row_page_max)+1
        row_current = (row-1)%self.row_page_max
        row_left = self.row_page_max - row_current

        return {"page": page_current, "row": row_current, "row_left": row_left}


    def new_page(self, row):
        """
        rowを次のページの1行目の行番号を返す
        """
        current_page = self.get_current_page(row)
        new_row = current_page["page"] * self.row_page_max + 1
        # print('NewPage', new_row, (new_row%self.row_page_max))
        return new_row
    

    def check_group_in_one_page(self, row, entry_num, head_row=2):
        # Param
        # - row: 書き始めの行番号
        # - entry_num: 書き込みエントリー数
        # - type: [ track, field ]

        current_page = self.get_current_page(row)
        row_need = head_row + entry_num*2

        if current_page["row_left"] > row_need:
            # 現在ページの空きに書き込める場合
            # print('ThisPage', row)
            return row, 'ThisPage'
        elif row_need < self.row_page_max:
            # １ページに1組全てを書き込める場合は改ページ
            # print('NewPage', row, (row%self.row_page_max+1))
            return self.new_page(row), 'NewPage'
        else:
            # 途中で改ページ
            # print('SplitPage', row)
            return row, 'SplitPage'

        
    
    def check_page(self, row, entry_num, type, title=True, ):
        # Param
        # - row: 書き始めの行番号
        # - entry_num: 書き込みエントリー数
        # - type: [ track, field ]
        row_page_max = 56       # 高さ13では56行で1ページ
        # Return
        # - 次に書き込む番号
        
        # 書き込みの最後の行の見積もり
        if type == 'track':
            row_end = entry_num + 2
        else:
            row_end = entry_num*2 + 2

        # 書き込むページ範囲を見積もり
        page_current = int(row/row_page_max)
        page_end = int(row_end/row_page_max)
        if row%row_page_max == 0: page_current = page_current - 1
        if row_end%row_page_max == 0: page_end = page_end - 1
        
        # ページをまたぐか否かの判断
        if page_current == page_end: # 改ページの必要なし
            return row
        elif page_current < page_end:
            return page_end * row_page_max

    
    """
    Write Title
    """
    def write_title_Track(self, ws, row, event, GR=None):
        # 種目名
        ws.merge_cells(start_row=row,start_column=1,end_row=row+2,end_column=3)
        self.write_cell(ws.cell(row=row, column=1), event.name)
        ws.cell(row=row, column=1).font = Font(size=28)
        # for i in range(3):
        #     for j in range(3):
        #         ws.cell(row=row+i, column=j+1).border = self.border_all_thick

        # 大会記録
        if GR:        
            self.write_cell(ws.cell(row=row+1, column=4), '大会記録')
            self.write_cell(ws.cell(row=row+1, column=5), format_mark(GR.mark, event)) #記録
            ws.merge_cells(start_row=row+1,start_column=6,end_row=row+1,end_column=7)
            self.write_cell(ws.cell(row=row+1, column=6), self.format_name(GR)) #氏名
            ws.merge_cells(start_row=row+1,start_column=8,end_row=row+1,end_column=9)
            self.write_cell(ws.cell(row=row+1, column=8), GR.club)#所属
            self.write_cell(ws.cell(row=row+1, column=10), GR.year) #年
        return row+3
        
        
    def write_title_HJPV(self, ws, row, event, GR=None):
        # 種目名
        ws.merge_cells(start_row=row,start_column=1,end_row=row+2,end_column=3)
        self.write_cell(ws.cell(row=row, column=1), event.name)
        ws.cell(row=row, column=1).font = Font(size=28)
        # for i in range(3):
        #     for j in range(3):
        #         ws.cell(row=row+i, column=j+1).border = self.border_all_thick
        # 大会記録
        if GR:
            self.write_cell(ws.cell(row=row+1, column=4), '大会記録')
            self.write_cell(ws.cell(row=row+1, column=5), format_mark(GR.mark, event)) #記録
            ws.merge_cells(start_row=row+1,start_column=6,end_row=row+1,end_column=11)
            self.write_cell(ws.cell(row=row+1, column=6), self.format_name(GR)) #氏名
            ws.merge_cells(start_row=row+1,start_column=12,end_row=row+1,end_column=16)
            self.write_cell(ws.cell(row=row+1, column=12), GR.club)#所属
            ws.merge_cells(start_row=row+1,start_column=17,end_row=row+1,end_column=19)
            self.write_cell(ws.cell(row=row+1, column=17), GR.year) #年            
        return row+3

    
    def write_title_LJTJThrow(self, ws, row, event, GR=None):
        # 種目名
        ws.merge_cells(start_row=row,start_column=1,end_row=row+2,end_column=3)
        self.write_cell(ws.cell(row=row, column=1), event.name)
        ws.cell(row=row, column=1).font = Font(size=24)
        # for i in range(3):
        #     for j in range(3):
        #         ws.cell(row=row+i, column=j+1).border = self.border_all_thick        
        # 大会記録
        if GR:
            self.write_cell(ws.cell(row=row+1, column=4), '大会記録')
            self.write_cell(ws.cell(row=row+1, column=5), format_mark(GR.mark,event)) #記録
            ws.merge_cells(start_row=row+1,start_column=6,end_row=row+1,end_column=7)
            self.write_cell(ws.cell(row=row+1, column=6), self.format_name(GR)) #氏名
            self.write_cell(ws.cell(row=row+1, column=8), GR.club)#所属
            self.write_cell(ws.cell(row=row+1, column=9), GR.year) #年                 
        return row+3


    def write_title(self, ws, row, event, GR=None):
        # 部門
        # GRのチェック
        try:
            GR = GRecord.objects.get(event=event, comp=self.comp)
        except GRecord.DoesNotExist:
            pass
        # エントリーを書き出し
        if event.program_type == 'Track8' or event.program_type == 'TrackN':
            row = self.write_title_Track(ws, row, event, GR)
        elif event.program_type == 'HJPV':
            row = self.write_title_HJPV(ws, row, event, GR)
        elif event.program_type == 'LJTJ' or event.program_type == 'Throw':
            row = self.write_title_LJTJThrow(ws, row, event, GR)       
        return row
    
    
    """
    Write Head
    """
    def write_head_Track(self, ws, row, col, group=True, wind=True):
        # 1行目
        if group and not group == -123:
            self.write_cell(ws.cell(row=row, column=col+1), str(group)+u"組", border=self.border_all)
        if wind:
            self.write_cell(ws.cell(row=row, column=col+5), u"(＋・ー", al=self.al_left) # Merged
            self.write_cell(ws.cell(row=row, column=col+6), u"m/s)", al=self.al_right) # Merged
        if group or wind: # どちらかを書き込む場合は
            row += 1
        
        # 2行目
        self.write_cell(ws.cell(row=row, column=col+1), u"ﾚｰﾝ", al=self.al_left)
        self.write_cell(ws.cell(row=row, column=col+2), u"No.", al=self.al_left)
        self.write_cell(ws.cell(row=row, column=col+3), u"氏名", al=self.al_left)
        self.write_cell(ws.cell(row=row, column=col+4), u"所属/陸協", al=self.al_left)
        self.write_cell(ws.cell(row=row, column=col+5), u"参考記録", al=self.al_left)
        self.write_cell(ws.cell(row=row, column=col+6), u"順位/記録", al=self.al_left)
        # 次の空白行の行番号を返す
        return row+1


    def write_head_Field(self, ws, row, group=None):
        # 1行目
        if group and not group == -123:
            self.write_cell(ws.cell(row=row, column=1), str(group)+u"組")
            row += 1
            
        # 2行目
        # 試技順
        ws.merge_cells(self.cell_posi(row, "A")+":"+self.cell_posi(row+1,"A"))
        self.write_cell(ws.cell(row=row, column=1), u"試順", al=self.al_wrap)
        # No, 学年
        self.write_cell(ws.cell(row=row, column=2), u"No.")
        self.write_cell(ws.cell(row=row+1, column=2), u"学年")
        # 氏名, フリガナ
        ws.merge_cells(self.cell_posi(row, "C")+":"+self.cell_posi(row+1,"C"))
        self.write_cell(ws.cell(row=row, column=3), u"氏名") # Merged
        # 所属, 陸協
        self.write_cell(ws.cell(row=row, column=4), u"所属")
        self.write_cell(ws.cell(row=row+1, column=4), u"陸協")
        # 参考記録
        ws.merge_cells(self.cell_posi(row, "E")+":"+self.cell_posi(row+1,"E"))
        self.write_cell(ws.cell(row=row, column=5), u"参考\n記録", al=self.al_wrap)
        return row
        
        
    def write_head_HJPV(self, ws, row, group=None):
        # Feild共通ヘッダー
        row = self.write_head_Field(ws,row, group=group)
        # 試技
        for i in np.arange(6,33, 3):
            ws.merge_cells(start_row=row,start_column=i,end_row=row+1,end_column=i+2)
            self.write_cell(ws.cell(row=row, column=i), "m")
            ws.cell(row=row, column=i).alignment =Alignment(vertical='bottom', horizontal='center')
        # 記録
        ws.merge_cells(self.cell_posi(row, "AG")+":"+self.cell_posi(row+1,"AG"))
        self.write_cell(ws.cell(row=row, column=33), u"記録")
        # 順位
        ws.merge_cells(self.cell_posi(row, "AH")+":"+self.cell_posi(row+1,"AH"))
        self.write_cell(ws.cell(row=row, column=34), u"順位", al=self.al_wrap)
    
        # Borderの設定
        for i in range(1,35):
            ws.cell(row=row, column=i).border = self.border_all
            ws.cell(row=row+1, column=i).border = self.border_all   
            
        return row+2

    
    def write_head_LJTJThrow(self, ws, row, trial=6, group=None):
        # Feild共通ヘッダー
        row = self.write_head_Field(ws,row, group=group)
        # 試技&記録
        if trial == 6:
            cells = ["1", "2", "3",  u"3回目\nベスト", "4","5", "6",  u"記録"]
        else:
            cells = ["1", "2", "3", u"記録"]
        for i in np.arange(6,6+len(cells),1):
            ws.merge_cells(start_row=row,start_column=i,end_row=row+1,end_column=i)
            self.write_cell(ws.cell(row=row, column=i), cells[i-6], al=self.al_wrap)
        # 順位
        if trial == 6:
            ws.merge_cells(self.cell_posi(row, "N")+":"+self.cell_posi(row+1,"N"))
            self.write_cell(ws.cell(row=row, column=14), u"順位", al=self.al_wrap)
        else:
            ws.merge_cells(self.cell_posi(row, "J")+":"+self.cell_posi(row+1,"J"))
            self.write_cell(ws.cell(row=row, column=10), u"順位", al=self.al_wrap)
        
        # Borderの設定
        if trial == 6: cell_end = 14
        else: cell_end = 10
        for i in range(1,cell_end+1):
            ws.cell(row=row, column=i).border = self.border_all
            ws.cell(row=row+1, column=i).border = self.border_all   
            
        return row+2

    
    """
    Write Row
    """
    def write_row_Track(self, ws, row, col, entry, lane):
        if entry.entry_status == 'DNS' or entry.check:
            self.write_cell(ws.cell(row=row, column=col+1), lane, font=self.font_red, al=self.al_left)
        else:
            self.write_cell(ws.cell(row=row, column=col+1), lane, al=self.al_left)
        self.write_cell(ws.cell(row=row, column=col+2), entry.bib, al=self.al_left)
        # 名前
        name_str = str(entry.name_family)+"\u3000"+str(entry.name_first)
        if not len(entry.grade) == 0: # 学年記入あり
            name_str += "("+str(entry.grade)+")"
        self.write_cell(ws.cell(row=row, column=col+3), name_str, al=self.al_left)
        self.write_cell(ws.cell(row=row+1, column=col+3), self.format_kana(entry), al=self.al_left)
        self.write_cell(ws.cell(row=row, column=col+4), entry.club, al=self.al_left)
        self.write_cell(ws.cell(row=row+1, column=col+4), entry.jaaf_branch, al=self.al_left)
        self.write_cell(ws.cell(row=row, column=col+5), format_mark(entry.personal_best, entry.event_status.event), al=self.al_left)
        # 記録欄
        self.write_cell(ws.cell(row=row, column=col+6), "(  )", al=self.al_left)
        if entry.entry_status == 'DNS':
            self.write_cell(ws.cell(row=row+1, column=col+6), "DNS", font=self.font_red)
        ws.cell(row=row+1, column=col+6).border = self.border_bottom
        return row+2
        


    def write_row_Field(self, ws,row, entry, order):
        # 試技順
        ws.merge_cells(self.cell_posi(row, "A")+":"+self.cell_posi(row+1,"A"))
        if entry.entry_status == 'DNS' or entry.check:
            self.write_cell(ws.cell(row=row, column=1), order, font=self.font_red)
        else:
            self.write_cell(ws.cell(row=row, column=1), order)
        # No, 学年
        self.write_cell(ws.cell(row=row, column=2), entry.bib)
        self.write_cell(ws.cell(row=row+1, column=2), entry.grade)
        # 氏名
        self.write_cell(ws.cell(row=row, column=3), self.format_name(entry))
        self.write_cell(ws.cell(row=row+1, column=3), self.format_kana(entry))
        # 所属, 陸協
        self.write_cell(ws.cell(row=row, column=4), entry.club)
        self.write_cell(ws.cell(row=row+1, column=4), entry.jaaf_branch)
        # 参考記録
        ws.merge_cells(self.cell_posi(row, "E")+":"+self.cell_posi(row+1,"E"))
        self.write_cell(ws.cell(row=row, column=5), format_mark(entry.personal_best, entry.event_status.event))
        return row

    
            
    def write_row_HJPV(self, ws, row, entry, order):
        # Field 共通ヘッダー
        row = self.write_row_Field(ws, row, entry, order)
        # 試技
        for i in np.arange(6,33, 1):
            ws.merge_cells(start_row=row,start_column=i,end_row=row+1,end_column=i)
        # 記録
        ws.merge_cells(self.cell_posi(row, "AG")+":"+self.cell_posi(row+1,"AG"))
        # 順位
        ws.merge_cells(self.cell_posi(row, "AH")+":"+self.cell_posi(row+1,"AH"))
        if entry.entry_status == 'DNS':
            self.write_cell(ws.cell(row=row, column=33), "DNS", font=self.font_red)
        
        # Borderの設定
        for i in range(1,35):
            ws.cell(row=row, column=i).border = self.border_all
            ws.cell(row=row+1, column=i).border = self.border_all
        return row+2


    def write_row_LJTJ(self, ws, row, entry, order, trial=6):
        # Field 共通ヘッダー
        row = self.write_row_Field(ws, row, entry, order)
        # 試技&記録
        if trial == 6: cell_end = 14
        else: cell_end = 10
        for i in np.arange(6,cell_end, 1):
            self.write_cell(ws.cell(row=row, column=i), "m", font=self.font_small)
            self.write_cell(ws.cell(row=row+1, column=i), "＋･－", font=self.font_small, al=self.al_left)        
        # 順位
        if trial == 6:
            ws.merge_cells(self.cell_posi(row, "N")+":"+self.cell_posi(row+1,"N"))
            if entry.entry_status == 'DNS':
                self.write_cell(ws.cell(row=row, column=14), "DNS", font=self.font_red)
        else:
            ws.merge_cells(self.cell_posi(row, "J")+":"+self.cell_posi(row+1,"J"))
            if entry.entry_status == 'DNS':
                self.write_cell(ws.cell(row=row, column=10), "DNS", font=self.font_red)
        
        # Borderの設定
        for i in range(1,cell_end+1):
            ws.cell(row=row, column=i).border = self.border_all
            ws.cell(row=row+1, column=i).border = self.border_all   
        return row + 2

    
    def write_row_Throw(self, ws, row, entry, order, trial=6):
        # Field 共通ヘッダー
        row = self.write_row_Field(ws, row, entry, order)
        # 試技&記録
        if trial == 6: cell_end = 14
        else: cell_end = 10
        for i in np.arange(6,cell_end, 1):
            ws.merge_cells(start_row=row,start_column=i,end_row=row+1,end_column=i)
            self.write_cell(ws.cell(row=row, column=i), "m", al=self.al_bottom)
        # 順位
        if trial == 6:
            ws.merge_cells(self.cell_posi(row, "N")+":"+self.cell_posi(row+1,"N"))
            if entry.entry_status == 'DNS':
                self.write_cell(ws.cell(row=row, column=14), "DNS", font=self.font_red)
        else:
            ws.merge_cells(self.cell_posi(row, "J")+":"+self.cell_posi(row+1,"J"))
            if entry.entry_status == 'DNS':
                self.write_cell(ws.cell(row=row, column=10), "DNS", font=self.font_red)
        
        # Borderの設定
        for i in range(1,cell_end+1):
            ws.cell(row=row, column=i).border = self.border_all
            ws.cell(row=row+1, column=i).border = self.border_all   
        return row + 2

    
    """
    Write Group
    """
    def write_group_lane8(self, ws, row,  entries, group=False, wind=True):
        # Params
        # - ws: worksheet
        # - row: 書き込み開始行番号
        # - entries: QuerySet of Entry Objects        
        
        # 右カラムか左カラムの決定
        try:
            if (group%2) == 0:      # 偶数組みは右カラムに
                col = 7
                row = self.row_prev_group_top
            else:
                col = 0
        except TypeError:
            col = 0
            
        # 同じページに全ての組みが書き込めるか確認
        current_page = self.get_current_page(row)
        if current_page["row_left"] < 18:
            # 8レーン全て書き込み不可の場合、改ページ
            row = self.new_page(row)
            col = 0
        elif current_page["row"] == 1:
            # ページの1行目が開いてる場合は詰める
            row = row - 1            
        # 書き込む行を保存
        self.row_prev_group_top = row
            
        
        # Header
        row = self.write_head_Track(ws, row, col, group, wind=entries[0].event_status.event.wind)
        # エントリーの書き込み
        c = 0
        for i in [0,1,2,3,4,5,6,7]:
            try:
                entry = entries.get(order_lane=i+1)
                c += 1
            except Entry.DoesNotExist:
                entry = self.entry_Null
            row = self.write_row_Track(ws, row, col, entry, i+1)
        # Finish
        print("> Write_group_lane8: write ", str(c),"/", len(entries), " entries")
        if row > self.row_prev_group_last:
            self.row_prev_group_last = row
            return row
        else:
            self.row_prev_group_last = row
            return self.row_prev_group_last


            
    # 長距離&補欠用
    def write_group_laneN(self, ws, row, entries, group=False, wind=True, col_right=False):
        # Params
        # - ws: worksheet
        # - row: 書き込み開始行番号
        # - entries: QuerySet of Entry Objects

        # 右カラムか左カラムの決定
        try:
            if col_right or (group%2) == 0:      # 偶数組と補欠は右カラムに
                col = 7
                row = self.row_prev_group_top
            else:
                col = 0
        except TypeError:
            col = 0
            
        # 同じページに全ての組みが書き込めるか確認
        current_page = self.get_current_page(row)
        max_lane = entries.aggregate(Max('order_lane')) # 最大のレーン
        row_need = 2 + max_lane["order_lane__max"]*2    # 書き込みに必要な行数の見積もり
        if current_page["row_left"] < row_need:
            # 8レーン全て書き込み不可の場合、改ページ
            row = self.new_page(row)
            col = 0
        elif current_page["row"] == 2:
            # ページの1行目が開いてる場合は詰める
            row = row - 1
        # 書き込む行を保存
        self.row_prev_group_top = row

        
        # Header
        row = self.write_head_Track(ws, row, col, group=group, wind=entries[0].event_status.event.wind)
        # エントリーの書き込み
        c = 0
        for i in np.arange(0, max_lane["order_lane__max"], 1):
            try:
                entry = entries.get(order_lane=i+1)
                c += 1
            except Entry.DoesNotExist:
                entry = self.entry_Null
            row = self.write_row_Track(ws, row, col, entry, i+1)
        # Finish
        print("> Write_group_laneN: write ", str(c),"/", len(entries), " entries")
        if row > self.row_prev_group_last:
            row_return = row
        else:
            row_return =  self.row_prev_group_last
        self.row_prev_group_last = row
        return row_return

    def write_group_HJPV(self, ws, row,  entries, group=False):
        # このページに全て書き込み可能か判定
        row, flg = self.check_group_in_one_page(row, len(entries))
        # Header
        row = self.write_head_HJPV(ws, row)
        # エントリーの書き込み
        c = 0
        if flg == 'SplitPage':  # ページ分割が必要な場合
            for entry in entries:
                current_page = self.get_current_page(row)
                if current_page["row_left"] < 2: # ページ分割
                    row = self.new_page(row)
                    row = self.write_head_HJPV(ws, row, trial=trial, group=group)
                row = self.write_row_HJPV(ws, row, entry, entry.order_lane, trial=trial)
                c +=1
        else:                   # ページ分割不要
            for entry in entries:
                row = self.write_row_HJPV(ws, row, entry, entry.order_lane)
                c += 1
        # Finish
        print("> Write_group_HJPV: write ", str(c),"/", len(entries), " entries")               
        return row


    def write_group_LJTJ(self, ws, row,  entries, trial=None, group=False):
        # このページに全て書き込み可能か判定
        row, flg = self.check_group_in_one_page(row, len(entries))
        # Trial
        if not trial:
            try:
                entry = entries.get(order_lane=1)
                if entry.event_status.section == 'VS': trial = 6
                else: trial = 3
            except ObjectDoesNotExist:
                pass
        # Header
        row = self.write_head_LJTJThrow(ws, row, trial=trial, group=group)
        # エントリーの書き込み
        c = 0
        if flg == 'SplitPage':  # ページ分割が必要な場合
            for entry in entries:
                current_page = self.get_current_page(row)
                if current_page["row_left"] < 2: # ページ分割
                    row = self.new_page(row)
                    row = self.write_head_LJTJThrow(ws, row, trial=trial, group=group)
                row = self.write_row_LJTJ(ws, row, entry, entry.order_lane, trial=trial)
                c +=1
        else:                   # ページ分割不要
            for entry in entries:
                row = self.write_row_LJTJ(ws, row, entry, entry.order_lane, trial=trial)
                c += 1
        # Finish
        print("> Write_group_LJTJ: write ", str(c),"/", len(entries), " entries")
        return row

            
    def write_group_Throw(self, ws, row,  entries, trial=6, group=False):
        # このページに全て書き込み可能か判定
        row, flg = self.check_group_in_one_page(row, len(entries))
        # Header
        row = self.write_head_LJTJThrow(ws, row, trial=trial, group=group)
        # エントリーの書き込み
        c = 0
        if flg == 'SplitPage':  # ページ分割が必要な場合
            for entry in entries:
                current_page = self.get_current_page(row)
                if current_page["row_left"] < 2: # ページ分割
                    row = self.new_page(row)
                    row = self.write_head_LJTJThrow(ws, row, trial=trial, group=group)
                row = self.write_row_Throw(ws, row, entry, entry.order_lane, trial=trial)
                c +=1
        else:
            for entry in entries:
                row = self.write_row_Throw(ws, row, entry, entry.order_lane, trial=trial)
                c +=1
        # Finish
        print("> Write_group_Throw: write ", str(c),"/", len(entries), " entries")               
        return row       

    
    def write_groups(self, ws, row, event, Entries):
        groups = Entries.values_list('group', flat=True).order_by('group').distinct()
        for group in groups:
            # 組数が負の場合
            if group < 0 and not group == -123 : # continue # 番組編成　未完了は書き出さない
                continue
            
            entries_g = Entries.filter(group=group).order_by('order_lane')
            # エントリーを書き出し
            if event.program_type == 'Track8':
                if group == -123: # 補欠は書き出し
                    row = self.write_group_laneN(ws, row, entries_g, group=group, wind=event.wind, col_right=True)
                else:
                    row = self.write_group_lane8(ws, row, entries_g, group=group, wind=event.wind)
            elif event.program_type == 'TrackN':
                if group == -123: # 補欠は書き出し
                    row = self.write_group_laneN(ws, row, entries_g, group=group, col_right=True)
                else:
                    row = self.write_group_laneN(ws, row, entries_g, group=group)
            elif event.program_type == 'HJPV':
                row = self.write_group_HJPV(ws, row, entries_g, group=group)
            elif event.program_type == 'LJTJ':
                row = self.write_group_LJTJ(ws, row, entries_g, group=group)
            elif event.program_type == 'Throw':
                row = self.write_group_Throw(ws, row, entries_g, group=group)
            # 1組書いた後は2行開ける
            row += 1
        return row

    
    """
    Sheet Styling
    """
    # Track
    def style_Track(self, ws, row, event=False):
        # Params
        # - ws: worksheet
        # - row: The last row number of the page you edit (type=int)
        # - col:  The last col number of the page you edit (type=int)
    
        #  列の幅の指定
        col_width = [("A", 3.0), ("B", 5.0), ("C", 11.0), ("D", 8.5), ("E", 6.0), ("F", 6.5),
                     ("G", 2.0),
                     ("H", 3.0), ("I", 5.0), ("J", 11.0), ("K", 8.5), ("L", 6.0), ("M", 6.5)]
        for t in col_width:
            ws.column_dimensions[t[0]].width = t[1]

        if event:
            ws.oddHeader.center.text = sex_to_ja(event.sex)+str(event.name)+"  &[Page] / &N"
            ws.oddHeader.center.size = 12
            ws.oddHeader.center.font = "ＭＳ ゴシック"
            ws.oddHeader.center.color = "000000"            

        return row

        
    def style_HJPV(self, ws, row, event=False):
        # Params
        # - ws: worksheet
        # - row: The last row number of the page you edit (type=int)
        # - col:  The last col number of the page you edit (type=int)
        
        #  列の幅の指定
        # [E-AF] = 1pt
        col_width = [("A", 2.5), ("B", 6.0), ("C", 10.5), ("D", 10.5), ("E", 4.5),  ("AG", 3.0),("AH", 3.0)]
        col_width_2 = 1.5
        for t in col_width:
            ws.column_dimensions[t[0]].width = t[1]
        for t in [chr(i) for i in range(65+5,65+26)]: # F - Z
            ws.column_dimensions[t].width = col_width_2
        for t in ["A"+chr(i) for i in range(65,65+6)]: #  AA- AF
            ws.column_dimensions[t].width = col_width_2

        if event:
            ws.oddHeader.center.text = sex_to_ja(event.sex)+str(event.name)+"  &[Page] / &N"
            ws.oddHeader.center.size = 12
            ws.oddHeader.center.font = "ＭＳ ゴシック"
            ws.oddHeader.center.color = "000000"
            
        return row
    
            
    def style_LJTJThrow(self, ws, row, event=False):
        # Params
        # - ws: worksheet
        # - row: The last row number of the page you edit (type=int)
        # - col:  The last col number of the page you edit (type=int)
        
        #  列の幅の指定
        # [E-AF] = 1pt
        col_width = [("A", 2.5), ("B", 6.0), ("C", 10.5), ("D", 10.0), ("E", 4.5),  ("N", 3.0)]
        col_width_2 = 5.5
        for t in col_width:
            ws.column_dimensions[t[0]].width = t[1]
        for t in [chr(i) for i in range(65+5,65+13)]: # F - M
            ws.column_dimensions[t].width = col_width_2


        # 印刷ヘッダーの編集
        if event:
            ws.oddHeader.center.text = sex_to_ja(event.sex)+str(event.name)+"  &[Page] / &N"
            ws.oddHeader.center.size = 12
            ws.oddHeader.center.font = "ＭＳ ゴシック"
            ws.oddHeader.center.color = "000000"

        return row


    def style_sheet(self, ws, row, event, mode=None):        
        #　行の高さの指定
        for  i in range(0,row):
            ws.row_dimensions[i].height = 13

        # 複数種目書き出しの場合
        if mode == 'many':
            event_cp = None
        else:
            event_cp = event
            
        # 種目ごとに分岐
        if event.program_type == 'Track8' or event.program_type == 'TrackN':
            return self.style_Track(ws, row, event=event_cp)
        elif event.program_type == 'HJPV':
            return self.style_HJPV(ws, row, event=event_cp)
        elif event.program_type == 'LJTJ' or event.program_type == 'Throw':
            return self.style_LJTJThrow(ws, row, event=event_cp)


        
    """
    Write Event
    """
    def write_class(self, ws, row, class_name, col=0):
        # 部門名を書き出し
        ws.merge_cells(start_row=row,start_column=col+1,end_row=row,end_column=col+3)
        self.write_cell(ws.cell(row=row, column=col+1), "< "+class_name+" >")
        ws.cell(row=row, column=col+1).alignment = self.al_left
        if col == 0:
            return row+1
        else:
            row_return = self.row_prev_group_last+1
            return row_return
        
    def write_event(self, ws, row, event, GR=None,VS=False,Sub=False,OP=False):
        # Params
        # - ws/row: Worksheet, 行番号
        # - event: Eventオブジェクト
        # - GR: 大会記録オブジェクト
        # - VS,SUb, OP: Entryオブジェクト(対校, 補欠, OP)

        # Entryがない場合
        if not VS and not Sub and not OP:
            return row
        
        # Title
        row = self.write_title(ws, row, event, GR)

        # 部門&レコード書き出し
        ## 対校
        if VS:
            row = self.write_class(ws, row, "対校")
            row = self.write_groups(ws, row, event, VS)            
        ## 補欠
        if Sub:
            if event.program_type == 'Track8' or event.program_type == 'TrackN':
                col = 7
                row = self.row_prev_group_top-1
            else:
                col = 0
            row = self.write_class(ws, row, "補欠", col=col)
            row = self.write_groups(ws, row, event, Sub)
        ## OP
        if OP:
            row = self.write_class(ws, row, "オープン")
            row = self.write_groups(ws, row, event, OP) 

        # 改ページ(次のページへ)
        return self.new_page(row)

    """
    Write Sheet through Django Cardinal System
    """
    def cardinal_write_by_event(self, ws, row, comp, event):
        # Params
        # - wb, row: Workbook, 書き始めの行番号
        # - comp: Comp Object
        # - event: Event Object

        # Entry
        ## 対校
        try:
            event_status = EventStatus.objects.filter(comp=comp, event=event, section='VS')
            VS = Entry.objects.filter(event_status=event_status, group__gt=0).order_by('group', 'order_lane')
        except (EventStatus.DoesNotExist, Entry.DoesNotExist) as e:
            VS = None
            print(e, "@cardinal_write_sheet_by_name")
        ## 補欠
        try:
            event_status = EventStatus.objects.filter(comp=comp, event=event, section='VS')
            Sub = Entry.objects.filter(event_status=event_status, group=-123).order_by('order_lane')
        except (EventStatus.DoesNotExist, Entry.DoesNotExist) as e:
            Sub = None
            print(e, "@cardinal_write_sheet_by_name")
        ## OP
        try:
            event_status = EventStatus.objects.filter(comp=comp, event=event, section='OP')
            OP = Entry.objects.filter(event_status=event_status, group__gt=0).order_by('group', 'order_lane')
        except (EventStatus.DoesNotExist, Entry.DoesNotExist) as e:
            OP = None
            print(e, "@cardinal_write_sheet_by_name")

        # 書き出し
        row = self.write_event(ws, row, event, VS=VS, Sub=Sub, OP=OP)
        return row
        

    def cardinal_write_by_event_status(self, ws, row, event_status):
        # Params
        # - wb, row: Workbook, 書き始めの行番号
        # - comp: Comp Object
        # - event: Event Object

        # 種目
        event = event_status.event

        # Entry
        try:
            entries = Entry.objects.filter(event_status=event_status).order_by('group', 'order_lane')
        except Entry.DoesNotExist as e:
            entries = None
            print(e, "@cardinal_write_sheet_by_name")
        # 部門
        if event_status.section == 'VS': VS = entries
        else: VS = None
        if event_status.section == 'Sub': Sub = entries
        else: Sub = None
        if event_status.section == 'OP': OP = entries
        else: OP = None
        
        # 書き出し
        row = self.write_event(ws, row, event, VS=VS, Sub=Sub, OP=OP)
        return row


    def cardinal_create_sheet_by_event(self, wb, comp, event, sheet_name=None):
        # Create Sheet
        ws = wb.create_sheet()
        row = 1
        if sheet_name: ws.title = sheet_name

        # 書き出し
        row = self.cardinal_write_by_event(ws, row, comp, event) 
        # シートのスタイリング
        row = self.style_sheet(ws, row, event)

        # シートの書き出し完了
        print("Success: write Excel Sheet of ", str(event))
        return True



    def cardinal_create_sheet_by_events(self, wb, comp, events, sheet_name=None):
        # 複数種目書き出し

        # Event Objectがない場合
        if len(events) == 0:
            return row;
        else:
            # Event Objectが同じProgram Typeを持つかチェック
            program_type = None
            for event in events:
                if program_type == 'Track' and (event.program_type == 'Track8' or event.program_type == 'TrackN'):
                    continue
                elif program_type == event.program_type:
                    continue
                elif program_type == None:
                    if (event.program_type == 'Track8' or event.program_type == 'TrackN'):
                        program_type = 'Track'
                    else:
                        program_type = event.program_type
                else:
                    print("Diffelent Program type: ", program_type, event.program_type)
                    return row

        # Create Sheet
        ws = wb.create_sheet()
        row = 1
        if sheet_name: ws.title = sheet_name
        # 書き出し
        for event in events:
            row = self.cardinal_write_by_event(ws, row, comp, event)
        # シートのスタイリング
        row = self.style_sheet(ws, row, event, mode='many')

        # シートの書き出し完了
        print("Success: write Excel Sheet of ", sheet_name)
        return True
    

    
    def cardinal_create_sheet_by_event_status(self, wb, event_status, sheet_name=None):
        # Create Sheet
        ws = wb.create_sheet()
        row = 1
        if sheet_name: ws.title = sheet_name

        # 書き出し
        event = event_status.event
        row = self.cardinal_write_by_event_status(ws, row, event_status) 
        # シートのスタイリング
        row = self.style_sheet(ws, row, event)

        # シートの書き出し完了
        print("Success: write Excel Sheet of ", str(event))
        return True    

    
    def cardinal_create_workbook_by_event(self, comp, event):
        # Create new Workbook
        wb = px.Workbook()
        ws = wb.active
        ws.title = 'None'

        #書き出し
        sheet_name = event.sex+event.name
        self.cardinal_create_sheet_by_event(wb, comp, event, sheet_name=sheet_name)

        ws = wb.get_sheet_by_name('None')
        wb.remove_sheet(ws)
        # 完了
        return wb
        
    
    
    def cardinal_create_workbook_by_event_status(self, comp, event_status):
        # Create new Workbook
        wb = px.Workbook()
        ws = wb.active
        ws.title = 'None'

        #書き出し
        event = event_status.event
        sheet_name = event_status.section + event.sex + event.name
        self.cardinal_create_sheet_by_event_status(wb, event_status, sheet_name=sheet_name)

        ws = wb.get_sheet_by_name('None')
        wb.remove_sheet(ws)
        # 完了
        return wb 
    

    def cardinal_create_workbook_field(self, comp, sex=None):
        # Params
        # - comp: Comp obkject
        # - sex: 性別
        
        # Create new Workbook
        wb = px.Workbook()
        ws = wb.active
        ws.title = 'None'

        #書き出し
        try:
            ## プログラムタイプでEventを取得
            events = Event.objects.filter(
                program_type__in=["HJPV","LJTJ","Throw"]
            ).order_by("sex", "start_list_priority")
            if sex:
                events = events.filter(sex=sex)
            ## シートの作成
            for event in events:
                sheet_name = event.sex+event.name
                self.cardinal_create_sheet_by_event(wb, comp, event, sheet_name=sheet_name)

        except Entry.DoesNotExist as e:
            print(e, "@cardinal_create_workbook_field")
            
        ws = wb.get_sheet_by_name('None')
        wb.remove_sheet(ws)
        # 完了
        return wb


    def cardinal_create_workbook_track(self, comp, mode='single'):
        # Params
        # - comp: Comp obkject
        # - sex: 性別
        # - mode: 'single'シート一枚に書き込む or 'multiple': 複数シートに書き込む
        
        # Create new Workbook
        wb = px.Workbook()
        ws = wb.active
        ws.title = 'None'

        #書き出し
        ## 男子
        try:
            ## プログラムタイプでEventを取得
            events = Event.objects.filter(
                program_type__in=["Track8","TrackN"],
                sex='M',
            ).order_by("start_list_priority", "name")
            ## シートの作成
            if mode == 'single':
                sheet_name = u"男トラック"
                self.cardinal_create_sheet_by_events(wb, comp, events, sheet_name=sheet_name)
            else:
                for event in events:
                    # 初期化
                    self.row_prev_group_top = 1
                    self.row_prev_group_last = 1
                    # 書き出し
                    sheet_name = event.sex+event.name
                    self.cardinal_create_sheet_by_event(wb, comp, event, sheet_name=sheet_name)
        except Entry.DoesNotExist as e:
            print(e, "@cardinal_create_workbook_field")

        ## 女子
        try:
            ## プログラムタイプでEventを取得
            events = Event.objects.filter(
                program_type__in=["Track8","TrackN"],
                sex='W',
            ).order_by("start_list_priority", "name")
            ## シートの作成
            if mode == 'single':
                sheet_name = u"女トラック"
                self.cardinal_create_sheet_by_events(wb, comp, events, sheet_name=sheet_name)
            else:
                for event in events:
                    # 初期化
                    self.row_prev_group_top = 1
                    self.row_prev_group_last = 1
                    # 書き出し
                    sheet_name = event.sex+event.name
                    self.cardinal_create_sheet_by_event(wb, comp, event, sheet_name=sheet_name)
        except Entry.DoesNotExist as e:
            print(e, "@cardinal_create_workbook_field")            
            
        ws = wb.get_sheet_by_name('None')
        wb.remove_sheet(ws)
        # 完了
        return wb
